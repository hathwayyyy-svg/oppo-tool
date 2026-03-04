# -*- coding: utf-8 -*-
import io
import re
import datetime as dt
from copy import copy
from pathlib import Path
from typing import Dict, Tuple, List, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ======================
# 固定配置
# ======================
SHEET_OUT = "引入产品详细信息"
SHEET_NEG_DEFAULT = "产品谈判记录表"

WRITE_FIELDS = {
    "品牌",
    "型号",
    "品类",
    "CPU型号",
    "网络制式",
    "摄像头",
    "屏幕",
    "电池",
    "预计采购票面价（元）",
    "预计零售价（元）",
    "合同预计数量（台）",
}

SUPPLIER_COL_S = 19  # S列

st.set_page_config(page_title="OPPO 引入回填", layout="wide")
st.title("OPPO 引入回填（上传2个文件 → 一键生成）")
st.caption("✅ 自动识别谈判表/入库表（顺序随意）｜✅ 数量按谈判表“同一行”硬绑定（不再文本匹配）｜✅ 删除多余模板行")


# ======================
# 工具函数
# ======================
def norm_text(x) -> str:
    if x is None:
        return ""
    return str(x).strip().replace("　", " ").replace("\u00a0", " ").strip()

def safe_set(cell, value):
    if isinstance(cell, MergedCell):
        return
    cell.value = value

def normalize_model_name(x) -> str:
    s = norm_text(x).upper()
    s = s.replace("（", "(").replace("）", ")")
    for bad in ["全网通", "移动", "联通", "电信", "公开版", "定制", "TD-LTE", "LTE", "NR"]:
        s = s.replace(bad, "")
    return s.strip()

def extract_model_token(s: str) -> str:
    s = normalize_model_name(s)
    m = re.search(r"[A-Z]{2,}\d+[A-Z0-9]*", s)
    return m.group(0) if m else s

def find_template_path() -> Path:
    for p in [Path("template.xlsx"), Path("template(1).xlsx"), Path("assets/template.xlsx")]:
        if p.exists():
            return p
    raise RuntimeError("仓库内未找到模板文件：template.xlsx / template(1).xlsx / assets/template.xlsx")


# ======================
# 自动识别文件类型
# ======================
def identify_excel_type(file_like) -> str:
    try:
        file_like.seek(0)
        wb = load_workbook(file_like, read_only=True, data_only=True)
        sheets = wb.sheetnames

        if SHEET_NEG_DEFAULT in sheets:
            return "negotiation"

        # 扫关键字段
        for s in sheets[:3]:
            ws = wb[s]
            found_quote = False
            found_model = False
            for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=80, values_only=True):
                for v in row:
                    if isinstance(v, str):
                        t = v.strip()
                        if "供应商报价（元/台）" in t:
                            found_quote = True
                        if t == "型号":
                            found_model = True
            if found_quote and found_model:
                return "negotiation"

        inbound_keys = ["CP型号", "电池容量", "屏幕尺寸", "主摄像头物理像素", "次摄像头物理像素"]
        for s in sheets[:3]:
            ws = wb[s]
            for r in range(1, min(81, ws.max_row + 1)):
                vals = []
                for c in range(1, min(ws.max_column, 200) + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.strip():
                        vals.append(v.strip())
                if not vals:
                    continue
                hits = sum(1 for k in inbound_keys if any(k in x for x in vals))
                if hits >= 2:
                    return "inbound"
    except Exception:
        pass
    return "unknown"

def split_two_files(files) -> Tuple:
    if len(files) != 2:
        raise RuntimeError("请一次上传 2 个文件：谈判记录表 + 入库资料信息表。")
    f1, f2 = files[0], files[1]
    t1, t2 = identify_excel_type(f1), identify_excel_type(f2)
    f1.seek(0); f2.seek(0)
    if t1 == "negotiation" and t2 == "inbound":
        return f1, f2
    if t2 == "negotiation" and t1 == "inbound":
        return f2, f1
    raise RuntimeError(f"识别失败：文件1={t1}, 文件2={t2}。请确认一个谈判表一个入库表。")


# ======================
# 谈判表：用 openpyxl 读取，绑定“行号” → 价格 + 数量（彻底避免匹配不上）
# ======================
def read_negotiation_with_rowid(neg_file_like) -> Tuple[pd.DataFrame, List[str], Dict[Tuple[int, str], int]]:
    """
    返回：
      df_items：每一行一个型号版本（含 __row__ 行号、__token__）
      suppliers：K-P 表头的6家公司
      qty_by_row_supplier：key=(谈判表行号, 供应商名) -> 数量
    """
    neg_file_like.seek(0)
    wb = load_workbook(neg_file_like, data_only=True)
    if SHEET_NEG_DEFAULT not in wb.sheetnames:
        raise RuntimeError("谈判表缺少 sheet：产品谈判记录表")
    ws = wb[SHEET_NEG_DEFAULT]

    # 找表头行（包含“型号”“供应商报价（元/台）”）
    header_row = None
    col_map = {}
    for r in range(1, 60):
        row_vals = [ws.cell(r, c).value for c in range(1, 120)]
        if any(norm_text(v) == "型号" for v in row_vals) and any(isinstance(v, str) and "供应商报价（元/台）" in v for v in row_vals if isinstance(v, str)):
            header_row = r
            for c in range(1, 120):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    col_map[v.strip()] = c
            break
    if not header_row:
        raise RuntimeError("谈判表找不到表头行（型号/供应商报价（元/台））。")

    # 供应商表头 K-P（你固定要求）
    supplier_cols = list(range(11, 17))  # K..P
    suppliers = [norm_text(ws.cell(header_row, c).value) for c in supplier_cols]
    suppliers = [s for s in suppliers if s]
    if not suppliers:
        raise RuntimeError("谈判表 K-P 表头没有识别到供应商名称。")

    # 必要列
    c_brand = col_map.get("品牌")
    c_model = col_map.get("型号")
    c_buy = col_map.get("供应商报价（元/台）")
    c_retail = col_map.get("零售价") or col_map.get("建议零售价")  # 容错
    if not (c_model and c_buy):
        raise RuntimeError("谈判表缺少必要列：型号 / 供应商报价（元/台）")

    rows = []
    qty_by_row_supplier: Dict[Tuple[int, str], int] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        model = ws.cell(r, c_model).value
        buy = ws.cell(r, c_buy).value
        if model is None or norm_text(model) == "":
            continue
        if buy is None or str(buy).strip() == "":
            continue  # 仍按“报价非空”为有效行

        brand = ws.cell(r, c_brand).value if c_brand else None
        retail = ws.cell(r, c_retail).value if c_retail else None

        rows.append({
            "品牌": brand,
            "型号": model,
            "供应商报价（元/台）": buy,
            "零售价": retail,
            "__row__": r,  # ✅ 行号
            "__token__": extract_model_token(str(model)),
        })

        # ✅ 数量：同一行 K-P 直接取，不需要任何匹配
        for idx, c in enumerate(supplier_cols):
            if idx >= len(suppliers):
                break
            supplier = suppliers[idx]
            v = ws.cell(r, c).value
            if v is None or str(v).strip() == "":
                continue
            try:
                qty_by_row_supplier[(r, supplier)] = int(float(v))
            except Exception:
                continue

    df_items = pd.DataFrame(rows)
    if df_items.empty:
        raise RuntimeError("谈判表未读取到有效型号行（请确认报价列有值）。")

    return df_items, suppliers, qty_by_row_supplier


# ======================
# 入库表：结构化表识别（按 token）
# ======================
def try_parse_inbound_as_table(inbound_file_like) -> Tuple[Dict[str, dict], List[dict]]:
    inbound_file_like.seek(0)
    wb = load_workbook(inbound_file_like, read_only=True, data_only=True)

    model_headers = ["型号", "终端型号", "产品型号", "机型", "终端型号/机型"]
    col_variants = {
        "cpu": ["CP型号", "CPU型号", "CPU"],
        "cam_main": ["主摄像头物理像素（万像素）", "主摄像头物理像素", "主摄像头像素（万像素）", "主摄像头像素"],
        "cam_sub": ["次摄像头物理像素（万像素）", "次摄像头物理像素", "副摄像头像素（万像素）", "次摄像头像素"],
        "screen": ["屏幕尺寸（英寸）", "屏幕尺寸", "屏幕尺寸(英寸)"],
        "battery": ["电池容量（mAH）", "电池容量", "电池容量(mAh)", "电池容量（mAh）"],
        "net": ["终端制式（TD-LTE/TD-SCDMA）", "终端制式", "网络制式", "制式"],
    }

    specs_map: Dict[str, dict] = {}
    debug_rows: List[dict] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        header_row = None
        header_values: Dict[str, int] = {}

        for r in range(1, min(81, ws.max_row + 1)):
            row_texts = {}
            for c in range(1, min(ws.max_column, 200) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    row_texts[v.strip()] = c

            model_col = None
            for mh in model_headers:
                if mh in row_texts:
                    model_col = row_texts[mh]
                    break
            if not model_col:
                continue

            found = {"model": model_col}
            hit = 0
            for key, names in col_variants.items():
                for nm in names:
                    if nm in row_texts:
                        found[key] = row_texts[nm]
                        hit += 1
                        break

            if hit >= 2:
                header_row = r
                header_values = found
                break

        if not header_row:
            continue

        mc = header_values["model"]
        for r in range(header_row + 1, ws.max_row + 1):
            mval = ws.cell(r, mc).value
            if mval is None or str(mval).strip() == "":
                continue
            token = extract_model_token(str(mval))

            sp = {}
            for key in ["cpu", "cam_main", "cam_sub", "screen", "battery", "net"]:
                c = header_values.get(key)
                sp[key] = ws.cell(r, c).value if c else None

            score = sum(1 for k in ["cpu", "screen", "battery", "cam_main", "cam_sub"] if sp.get(k) not in [None, ""])
            if score >= 2:
                old = specs_map.get(token)
                if not old:
                    specs_map[token] = sp
                else:
                    old_score = sum(1 for k in ["cpu", "screen", "battery", "cam_main", "cam_sub"] if old.get(k) not in [None, ""])
                    if score > old_score:
                        specs_map[token] = sp
                debug_rows.append({"sheet": sheet, "token": token, "score": score, **sp})

    return specs_map, debug_rows


def format_common_fields(specs: dict):
    cpu = specs.get("cpu")

    cam_main = specs.get("cam_main")
    cam_sub = specs.get("cam_sub")
    camera = ""
    if cam_main or cam_sub:
        main_txt = str(cam_main).strip() if cam_main is not None else ""
        sub_txt = str(cam_sub).strip() if cam_sub is not None else ""
        if main_txt and sub_txt:
            camera = f"主摄{main_txt}，次摄{sub_txt}"
        elif main_txt:
            camera = f"主摄{main_txt}"
        elif sub_txt:
            camera = f"次摄{sub_txt}"

    screen = specs.get("screen")
    screen_txt = f"{screen}英寸" if screen is not None and str(screen).strip() != "" else ""

    battery = specs.get("battery")
    battery_txt = str(battery).strip() if battery is not None and str(battery).strip() != "" else ""

    net_raw = specs.get("net")
    net_txt = ""
    if isinstance(net_raw, str):
        up = net_raw.upper()
        if "NR" in up or "5" in up:
            net_txt = "5G"
        elif "LTE" in up or "4" in up:
            net_txt = "4G"
        else:
            net_txt = net_raw.strip()

    return norm_text(cpu), camera, screen_txt, battery_txt, net_txt


# ======================
# 写入模板（关键：每一行都填，且删掉多余模板行）
# ======================
def fill_template(template_stream: io.BytesIO,
                  df_items: pd.DataFrame,
                  suppliers: List[str],
                  qty_by_row_supplier: Dict[Tuple[int, str], int],
                  specs_map: Dict[str, dict],
                  debug_rows: List[dict]) -> bytes:
    wb = load_workbook(template_stream)
    ws = wb[SHEET_OUT]

    # 找表头行
    header_row = None
    for r in range(1, 80):
        vals = [ws.cell(r, c).value for c in range(1, 180)]
        if "品牌" in vals and "型号" in vals and "CPU型号" in vals:
            header_row = r
            break
    if header_row is None:
        raise RuntimeError("模板找不到表头行（品牌/型号/CPU型号）。")

    header_to_col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            header_to_col[v.strip()] = c

    start_row = header_row + 1
    example_row = start_row

    if "型号" not in header_to_col:
        raise RuntimeError("模板缺少“型号”列。")
    model_col = header_to_col["型号"]

    per_item_rows = len(suppliers)
    total_needed_rows = len(df_items) * per_item_rows

    # 计算模板当前已有数据区行数（不要用“None”判断，模板可能是空字符串/公式）
    # 改用：向下找第一个“完全空行”（前10列都空）作为结束
    end = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        probe = [ws.cell(r, c).value for c in range(1, 11)]
        if all(v is None or str(v).strip() == "" for v in probe):
            break
        end = r
    existing = max(1, end - start_row + 1)

    # 不够行：插入并复制样式
    if existing < total_needed_rows:
        ws.insert_rows(start_row + existing, amount=total_needed_rows - existing)
        for i in range(existing, total_needed_rows):
            tgt_r = start_row + i
            for c in range(1, ws.max_column + 1):
                src = ws.cell(example_row, c)
                tgt = ws.cell(tgt_r, c)
                if isinstance(tgt, MergedCell):
                    continue
                tgt._style = copy(src._style)
                tgt.number_format = src.number_format
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.alignment = copy(src.alignment)
                tgt.protection = copy(src.protection)
                tgt.comment = None
            ws.row_dimensions[tgt_r].height = ws.row_dimensions[example_row].height

    # 多余行：直接删除（彻底消除 #DIV/0 的假象）
    if existing > total_needed_rows:
        ws.delete_rows(start_row + total_needed_rows, existing - total_needed_rows)

    def setv(r: int, header: str, value):
        if header not in WRITE_FIELDS:
            return
        c = header_to_col.get(header)
        if c:
            # 你要求“每一行单元格都填充”：这里用空字符串代替 None
            safe_set(ws.cell(r, c), "" if value is None else value)

    # 填充：谈判表每行 × 6公司
    for i, row in df_items.iterrows():
        model_raw = row.get("型号")
        brand = row.get("品牌")
        buy = row.get("供应商报价（元/台）")
        retail = row.get("零售价")
        token = norm_text(row.get("__token__"))
        row_id = int(row.get("__row__"))

        specs = specs_map.get(token, {})
        cpu, camera, screen_txt, battery_txt, net_txt = format_common_fields(specs)

        for j, supplier in enumerate(suppliers):
            r = start_row + i * per_item_rows + j

            safe_set(ws.cell(r, SUPPLIER_COL_S), supplier)

            setv(r, "品牌", brand)
            setv(r, "型号", model_raw)
            setv(r, "品类", "手机")

            setv(r, "CPU型号", cpu)
            setv(r, "网络制式", net_txt)
            setv(r, "摄像头", camera)
            setv(r, "屏幕", screen_txt)
            setv(r, "电池", battery_txt)

            # 价格：严格谈判表当前行
            if pd.notna(buy):
                setv(r, "预计采购票面价（元）", float(buy))
            else:
                setv(r, "预计采购票面价（元）", "")

            if pd.notna(retail):
                setv(r, "预计零售价（元）", float(retail))
            else:
                setv(r, "预计零售价（元）", "")

            # 数量：严格“同一行”K-P
            qty = qty_by_row_supplier.get((row_id, supplier))
            setv(r, "合同预计数量（台）", qty if qty is not None else "")

    # debug sheet：让你直接看到哪些 token 没识别到规格
    try:
        if "debug_入库识别" in wb.sheetnames:
            wb.remove(wb["debug_入库识别"])
        ws_dbg = wb.create_sheet("debug_入库识别")
        ws_dbg.append(["sheet", "token", "score", "cpu", "cam_main", "cam_sub", "screen", "battery", "net"])
        for d in debug_rows:
            ws_dbg.append([
                d.get("sheet"), d.get("token"), d.get("score"),
                d.get("cpu"), d.get("cam_main"), d.get("cam_sub"),
                d.get("screen"), d.get("battery"), d.get("net"),
            ])

        if "debug_未命中token" in wb.sheetnames:
            wb.remove(wb["debug_未命中token"])
        ws_miss = wb.create_sheet("debug_未命中token")
        ws_miss.append(["型号", "token", "是否在入库表识别到"])
        for _, row in df_items.iterrows():
            tk = norm_text(row.get("__token__"))
            ws_miss.append([row.get("型号"), tk, "YES" if tk in specs_map else "NO"])
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ======================
# UI
# ======================
uploaded_files = st.file_uploader(
    "上传 2 个 Excel（谈判记录表 + 入库资料信息表），顺序随意",
    type=["xlsx"],
    accept_multiple_files=True
)

run_btn = st.button("🚀 一键生成", type="primary")

if run_btn:
    try:
        if not uploaded_files or len(uploaded_files) != 2:
            st.warning("请一次上传 2 个文件")
            st.stop()

        neg_file, inbound_file = split_two_files(uploaded_files)
        st.info(f"识别结果：谈判表 = {neg_file.name} ｜ 入库表 = {inbound_file.name}")

        # ✅ 谈判表：一次性拿到（行号 + 价格 + 数量）
        df_items, suppliers, qty_by_row_supplier = read_negotiation_with_rowid(neg_file)

        # ✅ 入库表：规格 map
        specs_map, debug_rows = try_parse_inbound_as_table(inbound_file)
        if not specs_map:
            st.error("入库资料表没有识别到结构化规格表头（型号/CP型号/电池容量/屏幕尺寸等）。")
            st.stop()

        tpl_path = find_template_path()
        tpl_stream = io.BytesIO(tpl_path.read_bytes())

        out_bytes = fill_template(tpl_stream, df_items, suppliers, qty_by_row_supplier, specs_map, debug_rows)

        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"【生成】产品引入详细信息及风险评估_{ts}.xlsx"

        st.success("✅ 生成成功！请下载：")
        st.download_button(
            "⬇️ 下载结果文件",
            data=out_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with st.expander("核对信息（可选）"):
            st.write("谈判表有效行数（输出型号版本数）：", len(df_items))
            st.write("供应商数量（K-P）：", len(suppliers))
            st.write("入库表识别到 token 数：", len(specs_map))

    except Exception as e:
        st.error("运行失败（请看详细报错）")
        st.exception(e)
